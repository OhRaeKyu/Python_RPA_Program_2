# PPT 자동화 검증 프로그램 v01 
# - 자동화로 작업 된 파일 오류 검증 프로그램
# 작성자 : 오래규
# 작성일 : 2021-02-19

from openpyxl import load_workbook
import pyautogui
import pyperclip
import os

franchise = str(input("업체명을 입력하시오 : "))

file_title = "우편표지 취합본 " + franchise + ".pptx"

################## 엑셀에서 업체 정보 추출 ################## 
wb = load_workbook("로드뷰 조사 리스트.xlsx") # 엑셀 파일 로딩
ws = wb.get_sheet_by_name(franchise) # 엑셀 파일 시트명 입력

address = []  # 업체 주소
names = []  # 지점명
posts = []  # 우편번호

for y in range(2, ws.max_row + 1):
  # 엑셀에서 주소 추출
  ads = str(ws.cell(column = 4, row = y).value)
  address.append(ads)
  # 엑셀에서 지점명 추출
  name = str(ws.cell(column = 2, row = y).value + " " + ws.cell(column = 3, row = y).value)
  names.append(name)
  # 엑셀에서 우편번호 추출
  post = str(ws.cell(column = 6, row = y).value)
  posts.append(post)

################## PPT에서 작업 ##################
ppt_file = pyautogui.getWindowsWithTitle(file_title)[0] # 양식이 저장되어 있는 PPT 파일 로딩
ppt_file.activate()

pyautogui.press("Home") # PPT 첫 슬라이드로 이동
pyautogui.click(1000, 300)

pyautogui.sleep(0.5)

# 업무 시작
for i in range(ws.max_row - 1): 
  pyautogui.press("Tab") # 주소 좌표 이동
  pyautogui.press("F2")
  pyautogui.hotkey("Ctrl", "c")
  address_copieds = pyperclip.paste() # 클립보드에 복사 후 변수에 저장
  address_copied = address_copieds.split("\n")  # 행 바꿈을 기준으로 분할
  pyautogui.press("ESC")
  
  if address[i] not in address_copied[0]: # 주소 첫 번째 행 오류 검증
    print(str(i + 1) + " 번째 슬라이드에서 오류 발생")
    pyautogui.press("PageDown")
    continue
  elif names[i] not in address_copied[1]: # 주소 두 번째 행 오류 검증
    print(str(i + 1) + " 번째 슬라이드에서 오류 발생")
    pyautogui.press("PageDown")
    continue
    
  pyautogui.press("Tab") # 지점명 좌표 이동
  pyautogui.press("F2")
  pyautogui.hotkey("Ctrl", "c")
  names_copied = pyperclip.paste()
  pyautogui.press("ESC")
  
  if (names[i] + " 사장님 귀하") not in names_copied: # 지점명 오류 검증
    print(str(i + 1) + " 번째 슬라이드에서 오류 발생")
    pyautogui.press("PageDown")
    continue

  for j in range(8):
    pyautogui.press("Tab") # 우편번호 좌표 이동
  pyautogui.press("F2")
  pyautogui.hotkey("Ctrl", "c")
  posts_copied = pyperclip.paste()
  
  if posts_copied != posts[i]:  # 우편번호 오류 검증
    print(str(i + 1) + " 번째 슬라이드에서 오류 발생") 

  if i is not ws.max_row - 2:
    pyautogui.press("PageDown")

os.system("pause")